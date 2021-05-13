import json
import openpyxl
import unicodedata
import glob
import os

# known issues:
# 1
# it's hardcoded that the record values start at B4
# it was in every database so I didn't bother

# you can freely rename any of the dictionary value, not key, if you need to 
record_dictionary = {
    1 : "record_id",
    2 : "record_author",
    3 : "record_coauthor",
    4 : "record_editor",
    5 : "record_title",
    6 : "record_subtitle",
    7 : "record_original_edition",
    8 : "record_series",
    9 : "record_publication_date",
    10 : "record_edition",
    11 : "record_place_of_publication",
    12 : "record_publisher",
    13 : "record_source",
    14 : "record_volume",
    15 : "record_issue",
    16 : "record_pages",
    17 : "record_language",
    18 : "record_issn",
    19 : "record_doi",
    20 : "record_source_link",
    21 : "record_keywords",
    22 : "record_additional_info",
}

# creates new folder in current working directory
current_directory = os.getcwd()
new_directory = os.path.join(current_directory, r"database_json")
if not os.path.exists(new_directory):
    os.makedirs(new_directory)

# iterates for every xlsx file in current directory 
for database_file in glob.glob("*.xlsx"):
    print("Working on: " + database_file)
    # open file
    wb = openpyxl.load_workbook(database_file)
    
    # results
    result_dictionary = {}
    
    # iterates for every sheet in file
    # every record starts at B3
    # col and roww values represent that cell
    for sheet in wb.sheetnames:
        col = 2
        # available records
        available_records = {}
        temp_available_records = []
        while(wb[sheet].cell(row = 4, column = col).value != None):
            temp_available_records.append(wb[sheet].cell(row = 4, column = col).value)
            col += 1

        # to ensure that the sheet has valid data 
        pass_sheet = False
        for item in temp_available_records:
            if(type(item) is not int):
                pass_sheet = True
        # if not then take the next sheet
        if(pass_sheet):
            continue

        # local_id is the id that the record recives
        local_id = 0
        # roww is named like that because I was lazy
        roww = 5
        # iterates for every row (record)
        while(wb[sheet].cell(row = roww, column = 2).value != None):
            
            col = 2
            available_records_dict = {}
            # default values for every record
            for record in record_dictionary:
                if(record != 1):
                    available_records_dict[record_dictionary[record]] = "None"

            # iterates for every available record value
            for item in temp_available_records:
                # exceptions is a dictionary for every value that is spread on multiple columns
                exceptions = {}
                exception_id = 0
                # if record value already exists in currently reading main record
                if(record_dictionary[item] in available_records_dict):
                    # if it's already dictionary, json
                    # it will only save data that is not empty, e.g. None
                    if(type(available_records_dict[record_dictionary[item]]) == dict):
                        exceptions = available_records_dict[record_dictionary[item]]
                        for saved_except in available_records_dict[record_dictionary[item]]:
                            if(exceptions[saved_except] != None and exceptions[saved_except] != "None"):
                                exceptions[saved_except] = exceptions[saved_except]
                                exception_id += 1
                        if(str(wb[sheet].cell(row = roww, column = col).value) != None and str(wb[sheet].cell(row = roww, column = col).value) != "None"):
                            exceptions[str(exception_id)] = str(wb[sheet].cell(row = roww, column = col).value)
                        # adds all of the multiple values (exceptions) to main record
                        available_records_dict[record_dictionary[item]] = exceptions

                    # if it's not empty and has a single value
                    elif(type(available_records_dict[record_dictionary[item]]) == str and available_records_dict[record_dictionary[item]] != None and available_records_dict[record_dictionary[item]] != "None"):
                        # if read value contains only spaces
                        if(str(wb[sheet].cell(row = roww, column = col).value).isspace()):
                            continue
                        # if read value is None then skip saving that record
                        elif(str(wb[sheet].cell(row = roww, column = col).value) == None or str(wb[sheet].cell(row = roww, column = col).value) == "None"):
                            continue
                        # adds to currenly saved value
                        else:
                            exceptions[str(0)] = available_records_dict[record_dictionary[item]]
                            exceptions[str(1)] = str(wb[sheet].cell(row = roww, column = col).value)
                            available_records_dict[record_dictionary[item]] = exceptions
                    
                    # if it's empty
                    elif(available_records_dict[record_dictionary[item]] == None or available_records_dict[record_dictionary[item]] == "None"):
                        # if read value contains only spaces
                        if(str(wb[sheet].cell(row = roww, column = col).value).isspace() == False):
                            available_records_dict[record_dictionary[item]] = str(wb[sheet].cell(row = roww, column = col).value)
                
                # next column
                col += 1 
            # next row
            roww += 1
            # add that main record to result json, its like: { id : {record json} }
            result_dictionary[local_id] = available_records_dict
            # next id
            local_id += 1

        # saves that json file with the same name as xlsx database
        with open("database_json/" + str(database_file.strip("xlsx") + "json"), 'w', encoding = "utf-8") as outfile:
            outfile.write(json.dumps(result_dictionary, sort_keys = False, indent = 6, ensure_ascii = False))